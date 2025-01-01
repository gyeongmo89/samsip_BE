from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from . import models
from .database import engine, get_db, Base
from pydantic import BaseModel
from typing import Optional, List
import logging

from sqlalchemy.orm import joinedload
from datetime import datetime
import io
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create database tables
Base.metadata.create_all(bind=engine)

# Create FastAPI app
app = FastAPI()

# CORS 설정
origins = [
    "http://localhost:3000",
    "http://localhost:8000",
    "https://samsip.vercel.app",
    "https://samsip-be.koyeb.app",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create database tables


@app.on_event("startup")
async def startup_event():
    logger.info("Creating database tables...")
    try:
        Base.metadata.create_all(bind=engine)
        logger.info("Database tables created successfully!")
    except Exception as e:
        logger.error(f"Error during startup: {e}")
        raise e
    logger.info("Database initialization completed!")


# Pydantic models


class SupplierBase(BaseModel):
    name: str
    contact: Optional[str] = None
    address: Optional[str] = None


class SupplierResponse(SupplierBase):
    id: int

    class Config:
        from_attributes = True


class ItemBase(BaseModel):
    name: str
    description: Optional[str] = None
    price: Optional[float] = None
    vat_excluded: Optional[bool] = None


class ItemCreate(ItemBase):
    pass


class ItemResponse(ItemBase):
    id: int

    class Config:
        from_attributes = True


class UnitBase(BaseModel):
    name: str
    description: Optional[str] = None


class UnitCreate(UnitBase):
    pass


class UnitResponse(UnitBase):
    id: int

    class Config:
        from_attributes = True


class OrderBase(BaseModel):
    supplier_id: int
    item_id: int
    unit_id: int
    quantity: float
    price: float
    total: float
    payment_schedule: str  # 대금지급주기
    purchase_cycle: str    # 구입주기
    client: str
    notes: Optional[str] = None
    date: Optional[str] = None


class OrderResponse(BaseModel):
    id: int
    date: Optional[str] = None
    supplier_id: int
    item_id: int
    unit_id: int
    quantity: float
    price: float
    total: float
    payment_schedule: str  # 대금지급주기
    purchase_cycle: str    # 구입주기
    client: str
    notes: Optional[str] = None
    supplier: SupplierResponse
    item: ItemResponse
    unit: UnitResponse
    approval_status: Optional[str] = None
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None
    rejection_reason: Optional[str] = None

    class Config:
        from_attributes = True


class OrderCreate(BaseModel):
    supplier_name: str
    item_name: str
    unit_name: str
    quantity: float
    price: float
    total: float
    payment_schedule: str  # 대금지급주기
    purchase_cycle: str    # 구입주기
    client: str
    notes: Optional[str] = None
    date: Optional[str] = None


class ApprovalRequest(BaseModel):
    password: str


class RejectionRequest(BaseModel):
    reason: str


def get_float_value(value):
    if not value:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if value.startswith("="):  # 엑셀 수식인 경우
            return 0.0  # 기본값 반환 또는 다른 처리 로직 추가
        try:
            return float(value.replace(",", ""))  # 천단위 구분자 제거
        except ValueError:
            return 0.0
    return 0.0


# API endpoints


@app.post("/suppliers/", response_model=SupplierResponse)
def create_supplier(supplier: SupplierBase, db: Session = Depends(get_db)):
    try:
        # Check if supplier with same name already exists
        existing_supplier = (
            db.query(models.Supplier)
            .filter(
                models.Supplier.name == supplier.name,
                # models.Supplier.is_deleted == False,
                models.Supplier.is_deleted is False,
            )
            .first()
        )
        if existing_supplier:
            raise HTTPException(status_code=400, detail="already_exists")

        db_supplier = models.Supplier(
            name=supplier.name,
            contact=supplier.contact,
            address=supplier.address,
            is_deleted=False,
        )
        db.add(db_supplier)
        db.commit()
        db.refresh(db_supplier)
        return db_supplier
    except Exception as e:
        db.rollback()
        if isinstance(e, HTTPException):
            raise e
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=400, detail="already_exists")
        logger.error(f"Error creating supplier: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/suppliers/", response_model=List[SupplierResponse])
def read_suppliers(db: Session = Depends(get_db)):
    suppliers = db.query(models.Supplier).all()
    return suppliers


@app.delete("/suppliers/bulk-delete")
def bulk_delete_suppliers(supplier_ids: List[int], db: Session = Depends(get_db)):
    try:
        # 실제로 데이터를 삭제
        db.query(models.Supplier).filter(models.Supplier.id.in_(supplier_ids)).delete(
            synchronize_session=False
        )
        db.commit()
        return {"message": "Suppliers deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/items/", response_model=ItemResponse)
def create_item(item: ItemCreate, db: Session = Depends(get_db)):
    try:
        # Check if item with same name already exists
        existing_item = (
            db.query(models.Item)
            .filter(
                models.Item.name == item.name,
                # models.Item.is_deleted == False,
                models.Item.is_deleted is False,
            )
            .first()
        )
        if existing_item:
            raise HTTPException(status_code=400, detail="already_exists")

        db_item = models.Item(
            name=item.name,
            description=item.description,
            price=item.price,
            is_deleted=False,
        )
        db.add(db_item)
        db.commit()
        db.refresh(db_item)
        return db_item
    except Exception as e:
        db.rollback()
        if isinstance(e, HTTPException):
            raise e
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=400, detail="already_exists")
        logger.error(f"Error creating item: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/items/", response_model=List[ItemResponse])
def read_items(db: Session = Depends(get_db)):
    items = db.query(models.Item).all()
    return items


@app.delete("/items/bulk-delete")
def bulk_delete_items(item_ids: List[int], db: Session = Depends(get_db)):
    try:
        db.query(models.Item).filter(models.Item.id.in_(item_ids)).delete(
            synchronize_session=False
        )
        db.commit()
        return {"message": "Items deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/units/", response_model=UnitResponse)
def create_unit(unit: UnitCreate, db: Session = Depends(get_db)):
    try:
        # Check if unit with same name already exists
        existing_unit = (
            db.query(models.Unit)
            .filter(
                models.Unit.name == unit.name,
                # models.Unit.is_deleted == False,
                models.Unit.is_deleted is False,
            )
            .first()
        )
        if existing_unit:
            raise HTTPException(status_code=400, detail="already_exists")

        db_unit = models.Unit(
            name=unit.name,
            description=unit.description,
            is_deleted=False,
        )
        db.add(db_unit)
        db.commit()
        db.refresh(db_unit)
        return db_unit
    except Exception as e:
        db.rollback()
        if isinstance(e, HTTPException):
            raise e
        if "UNIQUE constraint failed" in str(e):
            raise HTTPException(status_code=400, detail="already_exists")
        logger.error(f"Error creating unit: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/units/", response_model=List[UnitResponse])
def read_units(db: Session = Depends(get_db)):
    units = db.query(models.Unit).all()
    return units


@app.delete("/units/bulk-delete")
def bulk_delete_units(unit_ids: List[int], db: Session = Depends(get_db)):
    try:
        db.query(models.Unit).filter(models.Unit.id.in_(unit_ids)).delete(
            synchronize_session=False
        )
        db.commit()
        return {"message": "Units deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/orders/", response_model=List[OrderResponse])
def read_orders(db: Session = Depends(get_db)):
    try:
        orders = (
            db.query(models.Order)
            .filter(models.Order.is_deleted.is_(False))
            .options(
                joinedload(models.Order.supplier),
                joinedload(models.Order.item),
                joinedload(models.Order.unit),
            )
            .all()
        )

        # 삭제된 데이터에 대한 처리
        for order in orders:
            if not order.supplier or order.supplier.is_deleted:
                order.supplier = models.Supplier(id=0, name="삭제됨")
            if not order.item or order.item.is_deleted:
                order.item = models.Item(id=0, name="삭제됨")
            if not order.unit or order.unit.is_deleted:
                order.unit = models.Unit(id=0, name="삭제됨")

        return orders
    except Exception as e:
        logger.error(f"Error reading orders: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/orders/", response_model=OrderResponse)
def create_order(order: OrderBase, db: Session = Depends(get_db)):
    try:
        db_order = models.Order(
            supplier_id=order.supplier_id,
            item_id=order.item_id,
            unit_id=order.unit_id,
            quantity=order.quantity,
            price=order.price,
            total=order.total,
            payment_schedule=order.payment_schedule,
            purchase_cycle=order.purchase_cycle,
            client=order.client,
            notes=order.notes,
            date=order.date,
        )
        db.add(db_order)
        db.commit()
        db.refresh(db_order)
        return db_order
    except Exception as e:
        logger.error(f"Error creating order: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/orders/{order_id}")
def update_order(order_id: int, order: OrderCreate, db: Session = Depends(get_db)):
    db_order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not db_order:
        raise HTTPException(status_code=404, detail="Order not found")

    # 구입처 찾기 또는 생성
    supplier = (
        db.query(models.Supplier)
        .filter(models.Supplier.name == order.supplier_name)
        .first()
    )
    if not supplier:
        supplier = models.Supplier(name=order.supplier_name)
        db.add(supplier)
        db.flush()

    # 품목 찾기 또는 생성
    item = db.query(models.Item).filter(models.Item.name == order.item_name).first()
    if not item:
        item = models.Item(name=order.item_name)
        db.add(item)
        db.flush()

    # 단위 찾기 또는 생성
    unit = db.query(models.Unit).filter(models.Unit.name == order.unit_name).first()
    if not unit:
        unit = models.Unit(name=order.unit_name)
        db.add(unit)
        db.flush()

    # 발주 데이터 업데이트
    db_order.quantity = order.quantity
    db_order.price = order.price
    db_order.total = order.total
    db_order.payment_schedule = order.payment_schedule  # 대금지급주기
    db_order.purchase_cycle = order.purchase_cycle      # 구입주기
    db_order.client = order.client
    db_order.notes = order.notes
    db_order.date = order.date

    db_order.supplier_id = supplier.id
    db_order.item_id = item.id
    db_order.unit_id = unit.id

    db.commit()
    db.refresh(db_order)
    return db_order


@app.delete("/orders/bulk-delete")
def delete_orders(ids: List[int], db: Session = Depends(get_db)):
    try:
        for order_id in ids:
            order = db.query(models.Order).filter(models.Order.id == order_id).first()
            if order:
                db.delete(order)
        db.commit()
        return {"message": "Orders deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/orders/upload")
async def upload_orders(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    try:
        contents = await file.read()
        wb = load_workbook(
            filename=io.BytesIO(contents), data_only=True
        )  # data_only=True로 수식 대신 값을 가져옴
        ws = wb.active

        orders = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 빈 행 건너뛰기
                continue

            try:
                date_str = str(row[0])
                order_date = None

                if isinstance(row[0], datetime):  # Excel datetime 객체인 경우
                    order_date = row[0].date()
                elif "." in date_str:
                    try:
                        # 날짜 문자열에서 공백 제거 및 분리
                        parts = date_str.strip().split(".")
                        if len(parts) == 3:
                            year = int(parts[0])
                            month = int(parts[1])
                            day = int(parts[2])
                            order_date = datetime(year, month, day).date()
                    except (ValueError, TypeError, IndexError) as e:
                        print(f"Error parsing date with parts '{date_str}': {str(e)}")

                if order_date is None:
                    # 다른 형식들 시도
                    date_formats = ["%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"]
                    for date_format in date_formats:
                        try:
                            order_date = datetime.strptime(
                                date_str.strip(), date_format
                            ).date()
                            break
                        except ValueError:
                            continue

                if order_date is None:
                    print(f"Using current date for invalid date: '{date_str}'")
                    order_date = datetime.now().date()

            except Exception as e:
                print(f"Error processing date '{str(row[0])}': {str(e)}")
                order_date = datetime.now().date()

            # 구입처 찾기 또는 생성
            supplier = (
                db.query(models.Supplier).filter(models.Supplier.name == row[1]).first()
            )
            if not supplier:
                supplier = models.Supplier(name=row[1], contact=row[9])
                db.add(supplier)
                db.flush()

            # 품목 찾기 또는 생성
            item = db.query(models.Item).filter(models.Item.name == row[2]).first()
            if not item:
                notes = str(row[10] or "")
                vat_excluded = "부가세별도" in notes
                item = models.Item(
                    name=row[2],
                    price=get_float_value(row[3]),
                    vat_excluded=vat_excluded,
                    description="부가세별도" if vat_excluded else None,
                )
                db.add(item)
                db.flush()

            # 단위 찾기 또는 생성
            unit_name = str(row[4] or "개")  # None이면 기본값 "개" 사용
            unit = db.query(models.Unit).filter(models.Unit.name == unit_name).first()
            if not unit:
                unit = models.Unit(name=unit_name)
                db.add(unit)
                db.flush()

            # 발주 데이터 생성
            order = models.Order(
                date=order_date,
                supplier_id=supplier.id,
                item_id=item.id,
                unit_id=unit.id,
                price=get_float_value(row[3]),
                quantity=get_float_value(row[5]),
                total=get_float_value(row[6]),
                payment_schedule=str(row[7] or "미정"),
                purchase_cycle=str(row[8] or "daily"),
                client=str(row[9] or ""),
                notes=str(row[10] or ""),
            )
            orders.append(order)

        db.add_all(orders)
        db.commit()

        return {"message": f"Successfully uploaded {len(orders)} orders"}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/suppliers/{supplier_id}", response_model=SupplierResponse)
def update_supplier(
    supplier_id: int, supplier: SupplierBase, db: Session = Depends(get_db)
):
    db_supplier = (
        db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    )
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    for key, value in supplier.dict().items():
        setattr(db_supplier, key, value)

    db.commit()
    db.refresh(db_supplier)
    return db_supplier


@app.put("/items/{item_id}", response_model=ItemResponse)
def update_item(item_id: int, item: ItemBase, db: Session = Depends(get_db)):
    db_item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")

    for key, value in item.dict().items():
        setattr(db_item, key, value)

    db.commit()
    db.refresh(db_item)
    return db_item


@app.put("/units/{unit_id}", response_model=UnitResponse)
def update_unit(unit_id: int, unit: UnitCreate, db: Session = Depends(get_db)):
    try:
        db_unit = db.query(models.Unit).filter(models.Unit.id == unit_id).first()
        if not db_unit:
            raise HTTPException(status_code=404, detail="Unit not found")

        # Check if another unit with the same name exists
        existing_unit = (
            db.query(models.Unit)
            .filter(
                models.Unit.name == unit.name,
                models.Unit.id != unit_id,
                models.Unit.is_deleted is False,
            )
            .first()
        )
        if existing_unit:
            raise HTTPException(status_code=400, detail="already_exists")

        # Convert None to empty string for description
        unit_data = unit.dict()
        if unit_data.get("description") is None:
            unit_data["description"] = ""

        for key, value in unit_data.items():
            setattr(db_unit, key, value)

        db.commit()
        db.refresh(db_unit)
        return db_unit
    except HTTPException as e:
        raise e
    except Exception as e:
        db.rollback()
        logger.error(f"Error updating unit: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/orders/{order_id}/approve")
def approve_order(
    order_id: int, approval: ApprovalRequest, db: Session = Depends(get_db)
):
    if approval.password != "admin":
        raise HTTPException(status_code=401, detail="Invalid password")

    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order.approval_status = "approved"
    order.approved_by = "이지은"
    order.approved_at = datetime.now().strftime("%y-%m-%d %H:%M")

    db.commit()
    return {"message": "Order approved successfully"}


@app.post("/orders/{order_id}/reject")
def reject_order(
    order_id: int, rejection: RejectionRequest, db: Session = Depends(get_db)
):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    order.approval_status = "rejected"
    order.rejection_reason = rejection.reason
    order.approved_at = datetime.now().strftime("%y-%m-%d %H:%M")

    db.commit()
    return {"message": "Order rejected successfully"}
